# app.py

from flask import Flask, request, send_file
from openpyxl import load_workbook
from flask_cors import CORS  # Importar CORS
import os

app = Flask(__name__)
CORS(app)  # Habilitar CORS para todas las rutas

# Ruta del archivo de plantilla de Excel
TEMPLATE_PATH = 'template/PREOPERACIONALES.xlsx'  # Ruta ajustada
OUTPUT_PATH = 'plantilla_modificada.xlsx'

@app.route('/rellenar_excel', methods=['POST'])
def rellenar_excel():
    # Cargar datos del request JSON
    data = request.json

    try:
        # Cargar el archivo de plantilla de Excel
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active  # Obtener la hoja activa (puedes cambiar esto según tu necesidad)

        # Buscar la celda fusionada que contiene "item"
        fila_items = None
        for merged_range in ws.merged_cells.ranges:
            # Obtener la celda superior izquierda de cada rango fusionado
            min_row, min_col = merged_range.min_row, merged_range.min_col
            celda = ws.cell(row=min_row, column=min_col)
            if celda.value and str(celda.value).strip().lower() == "item":
                fila_items = min_row  # Guardar la fila donde se encuentra "item"
                break

        if not fila_items:
            return "No se encontró la fila 'item' en la plantilla.", 404

        # Determinar dinámicamente las columnas de los días de la semana a partir de la fila "item"
        dias_columna = {}
        for col in range(1, ws.max_column + 1):
            celda = ws.cell(row=fila_items, column=col)
            dia = celda.value
            if dia and dia.strip().lower() in ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado', 'domingo']:
                columna_true = ws.cell(row=fila_items+1, column=col).column_letter
                columna_false = ws.cell(row=fila_items+1, column=col+1).column_letter
                dias_columna[dia.strip().capitalize()] = (columna_true, columna_false)

        # Función para verificar si una celda está fusionada
        def obtener_celda_principal(hoja, celda):
            for merged_range in hoja.merged_cells.ranges:
                if celda.coordinate in merged_range:
                    return hoja.cell(merged_range.min_row, merged_range.min_col)
            return celda

        # Iterar sobre las secciones del JSON
        for seccion, items in data.items():
            for item, dias in items.items():
                # Eliminar espacios en blanco y convertir a minúsculas los nombres de los items del JSON
                item_stripped_lower = item.strip().lower()

                # Encontrar la fila correspondiente al nombre del item
                for row in range(fila_items + 2, ws.max_row + 1):  # Ajustar según tu plantilla
                    # Eliminar espacios en blanco y convertir a minúsculas los nombres de los items en la plantilla
                    item_excel = ws[f'A{row}'].value
                    if item_excel and item_excel.strip().lower() == item_stripped_lower:  # Verificar si el nombre coincide
                        # Rellenar los días de la semana en sus respectivas columnas
                        for dia, valor in dias.items():
                            dia_key = dia.strip().capitalize()
                            if dia_key in dias_columna:
                                columna_true, columna_false = dias_columna[dia_key]
                                if valor == True:
                                    # Encontrar la celda para True
                                    celda_destino_true = ws[f"{columna_true}{row}"]
                                    celda_principal_true = obtener_celda_principal(ws, celda_destino_true)
                                    print(f'Item: "{item}", Día: "{dia}", Valor: True, Celda: {celda_principal_true.coordinate}')
                                    celda_principal_true.value = 'X'
                                elif valor == False:
                                    # Encontrar la celda para False
                                    celda_destino_false = ws[f"{columna_false}{row}"]
                                    celda_principal_false = obtener_celda_principal(ws, celda_destino_false)
                                    print(f'Item: "{item}", Día: "{dia}", Valor: False, Celda: {celda_principal_false.coordinate}')
                                    celda_principal_false.value = 'X'
                        break  # Salir del bucle una vez encontrado el item

        # Guardar el archivo modificado
        wb.save(OUTPUT_PATH)

        # Enviar el archivo de vuelta al cliente con el tipo de contenido adecuado
        return send_file(OUTPUT_PATH, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except FileNotFoundError:
        return "El archivo de plantilla de Excel no se encontró. Verifique la ruta.", 404

# Iniciar el servidor con Gunicorn
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get("PORT", 5000)))
