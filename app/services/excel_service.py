from openpyxl import load_workbook
from openpyxl.styles import Font
import os

# Construir la ruta relativa desde el directorio actual de ejecución
TEMPLATE_PATH = os.path.join(os.getcwd(), 'template', 'PREOPERACIONALES.xlsx')
OUTPUT_PATH = os.path.join(os.getcwd(), 'plantilla_modificada.xlsx')


def procesar_excel(data):
    # Verifica que el archivo de plantilla existe
    if not os.path.exists(TEMPLATE_PATH):
        print(f"ERROR: No se encontró el archivo de plantilla en {TEMPLATE_PATH}")
        raise FileNotFoundError("El archivo de plantilla de Excel no se encontró. Verifique la ruta.")

    # cargar el archivo de plantilla de Excel
    try:
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active
    except Exception as e:
        print(f"ERROR al cargar el archivo de plantilla: {str(e)}")
        raise e

    # Extraer el objeto FORMULARIO del JSON y eliminarlo del objeto original
    formulario_data = data.pop('FORMULARIO', None)
    pie_tabla = data.pop("PIE_TABLA", None)

    # Llamar a la función que llena el formulario si existe
    if formulario_data:
        rellenar_formulario(ws, formulario_data)

    if pie_tabla:
        rellenar_pie_tabla(ws, pie_tabla)

    # Llamar a la función que llena la tabla en la sección específica
    rellenar_tabla(ws, data)

    # Guardar el archivo modificado
    wb.save(OUTPUT_PATH)


def obtener_rango_fusionado(hoja, celda):
    # Verifica si una celda está fusionada y obtiene la celda superior izquierda
    for merged_range in hoja.merged_cells.ranges:
        if celda.coordinate in merged_range:
            return merged_range, hoja.cell(merged_range.min_row, merged_range.min_col)
    return None, celda



def rellenar_formulario(ws, data):
    codigo = data.pop("Codigo", None)
    fecha_emision = data.pop("Fecha de Emision", None)

    # Función para limpiar y normalizar el texto
    def normalizar_texto(texto):
        if isinstance(texto, str):  # Verifica si es una cadena de texto
            return texto.strip().lower().rstrip(':')
        return str(texto)  # Convierte otros tipos de datos a string

    # Función para actualizar el contenido de la celda sin cambiar la etiqueta y aplicar formato
    def actualizar_celda_con_etiqueta(celda, etiqueta, nuevo_valor):
        contenido_actual = str(celda.value)
        if etiqueta.lower() in contenido_actual.lower():
            # Mantener la etiqueta y reemplazar solo el valor
            partes = contenido_actual.split(':', 1)  # Separar por el primer ':'
            if len(partes) > 1:
                # Mantener la etiqueta original y aplicar negrita a la etiqueta
                nuevo_contenido = f"{partes[0].strip()}: {nuevo_valor}"
                celda.value = nuevo_contenido

                # Aplicar negrita a la etiqueta y normal al valor
                negrita_fuente = Font(bold=True)
                normal_fuente = Font(bold=False)

                # Aplicar estilos a la etiqueta y al valor
                celda.font = negrita_fuente  # Aplicar negrita a la celda completa (necesario para inicializar)
                partes_etiqueta = len(partes[0]) + 1  # Longitud de la etiqueta más ":"
                for i in range(partes_etiqueta, len(nuevo_contenido)):
                    celda.font = normal_fuente  # Cambiar el resto a fuente normal

            else:
                print(f"No se pudo actualizar la celda {celda.coordinate} correctamente.")
        else:
            print(f"No se encontró la etiqueta '{etiqueta}' en la celda {celda.coordinate}.")

    # Buscar y actualizar el "Codigo"
    if codigo:
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value and "codigo:" in str(cell.value).strip().lower():
                    # Obtener la celda principal en caso de ser fusionada
                    rango_fusionado, celda_principal = obtener_rango_fusionado(ws, cell)
                    # Actualizar el valor en la celda encontrada
                    actualizar_celda_con_etiqueta(celda_principal, "Codigo", codigo)
                    found = True
                    break
            if found:
                break
        if not found:
            print("No se encontró una celda para 'Codigo'.")

    # Buscar y actualizar la "Fecha de Emision"
    if fecha_emision:
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value and "fecha de emision:" in str(cell.value).strip().lower():
                    # Obtener la celda principal en caso de ser fusionada
                    rango_fusionado, celda_principal = obtener_rango_fusionado(ws, cell)
                    # Actualizar el valor en la celda encontrada
                    actualizar_celda_con_etiqueta(celda_principal, "Fecha de Emision", fecha_emision)
                    found = True
                    break
            if found:
                break
        if not found:
            print("No se encontró una celda para 'Fecha de Emision'.")

    # Iterar sobre los datos del formulario
    for key, value in data.items():
        # Normalizar la clave del JSON
        key_normalizado = normalizar_texto(key)
        found = False
        # Buscar la celda que coincide con la clave normalizada
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                # Normalizar el valor de la celda
                cell_value_normalizado = normalizar_texto(cell.value)
                # Si el valor de la celda coincide con la clave del formulario
                if cell_value_normalizado == key_normalizado:
                    # Obtener el rango fusionado y la celda principal
                    rango_fusionado, celda_principal = obtener_rango_fusionado(ws, cell)

                    if rango_fusionado:
                        # Determinar la celda a la derecha del rango fusionado
                        col_final = rango_fusionado.max_col + 1
                    else:
                        # Si no está fusionada, solo se mueve a la derecha
                        col_final = cell.column + 1

                    # Asignar el valor en la celda después del rango fusionado
                    ws.cell(row=cell.row, column=col_final).value = value

                    found = True
                    break
            if found:
                break
        if not found:
            print(f"No se encontró una celda para la clave '{key}'.")

def rellenar_tabla(ws, data):
    # Buscar la celda fusionada que contiene "item"
    fila_items = None
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col = merged_range.min_row, merged_range.min_col
        celda = ws.cell(row=min_row, column=min_col)
        if celda.value and str(celda.value).strip().lower() == "item":
            fila_items = min_row  # Guardar la fila donde se encuentra "item"
            break

    if not fila_items:
        raise ValueError("No se encontró la fila 'item' en la plantilla.")

    # Determinar dinámicamente las columnas de los días de la semana a partir de la fila "item"
    dias_columna = {}
    for col in range(1, ws.max_column + 1):
        celda = ws.cell(row=fila_items, column=col)
        dia = celda.value
        if dia and dia.strip().lower() in ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado', 'domingo']:
            columna_true = ws.cell(row=fila_items + 1, column=col).column_letter
            columna_false = ws.cell(row=fila_items + 1, column=col + 1).column_letter
            dias_columna[dia.strip().capitalize()] = (columna_true, columna_false)

    # Función para verificar si una celda está fusionada
    def obtener_celda_principal(hoja, celda):
        for merged_range in hoja.merged_cells.ranges:
            if celda.coordinate in merged_range:
                return hoja.cell(merged_range.min_row, merged_range.min_col)
        return celda

    # Iterar sobre las secciones del JSON
    for seccion, items in data.items():
        for item, dias in items.items():
            # Eliminar espacios en blanco y convertir a minúsculas los nombres de los items del JSON
            item_stripped_lower = item.strip().lower()

            # Encontrar la fila correspondiente al nombre del item
            for row in range(fila_items + 2, ws.max_row + 1):
                # Eliminar espacios en blanco y convertir a minúsculas los nombres de los items en la plantilla
                item_excel = ws[f'A{row}'].value
                if item_excel and item_excel.strip().lower() == item_stripped_lower:
                    # Rellenar los días de la semana en sus respectivas columnas
                    for dia, valor in dias.items():
                        dia_key = dia.strip().capitalize()
                        if dia_key in dias_columna:
                            columna_true, columna_false = dias_columna[dia_key]
                            if valor == True:
                                celda_destino_true = ws[f"{columna_true}{row}"]
                                celda_principal_true = obtener_celda_principal(ws, celda_destino_true)
                                celda_principal_true.value = 'X'
                            elif valor == False:
                                celda_destino_false = ws[f"{columna_false}{row}"]
                                celda_principal_false = obtener_celda_principal(ws, celda_destino_false)
                                celda_principal_false.value = 'X'
                    break


def rellenar_pie_tabla(ws, data):
    nombre_conductor = data.pop("Nombre del Conductor", None)
    observaciones = data.pop("OBSERVACIONES", None)

    # Llenar el nombre del conductor en la celda anterior
    if nombre_conductor:
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                # Normalizar el valor de la celda
                if cell.value and "nombre del conductor" in str(cell.value).strip().lower():
                    # Obtener el rango fusionado y la celda principal
                    rango_fusionado, celda_principal = obtener_rango_fusionado(ws, cell)

                    if rango_fusionado:
                        # Determinar la celda antes del rango fusionado
                        col_anterior = rango_fusionado.min_col - 1

                        if col_anterior >= 1:
                            celda_destino = ws.cell(row=cell.row, column=col_anterior)
                            celda_principal_destino = obtener_rango_fusionado(ws, celda_destino)[1]
                            celda_principal_destino.value = nombre_conductor

                    else:
                        # Si no está fusionada, simplemente toma la celda anterior
                        col_anterior = cell.column - 1
                        if col_anterior >= 1:
                            ws.cell(row=cell.row, column=col_anterior).value = nombre_conductor

                    found = True
                    break
            if found:
                break
        if not found:
            print("No se encontró una celda para 'Nombre del Conductor'.")

    # Llenar las observaciones en la celda debajo de la identificada
    if observaciones:
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                # Normalizar el valor de la celda
                if cell.value and "observaciones" in str(cell.value).strip().lower():
                    # Obtener el rango fusionado y la celda principal
                    rango_fusionado, celda_principal = obtener_rango_fusionado(ws, cell)

                    if rango_fusionado:
                        # Determinar la celda debajo del rango fusionado
                        fila_inferior = rango_fusionado.max_row + 1

                        if fila_inferior <= ws.max_row:
                            celda_destino = ws.cell(row=fila_inferior, column=cell.column)
                            celda_principal_destino = obtener_rango_fusionado(ws, celda_destino)[1]
                            celda_principal_destino.value = observaciones

                    else:
                        # Si no está fusionada, simplemente toma la celda de abajo
                        fila_inferior = cell.row + 1
                        if fila_inferior <= ws.max_row:
                            ws.cell(row=fila_inferior, column=cell.column).value = observaciones

                    found = True
                    break
            if found:
                break
        if not found:
            print("No se encontró una celda para 'OBSERVACIONES'.")



