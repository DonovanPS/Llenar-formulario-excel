from datetime import datetime, timedelta
import io
import os
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as XLImage
import requests
from io import BytesIO

def get_template_path():
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
    return os.path.join(base_dir, 'template', 'LIMPIEZA.xlsx')

def validar_datos_inspeccion(inspeccion_data):
    """Valida la estructura y valores de los datos de inspección"""
    for elemento, dias in inspeccion_data.items():
        if not isinstance(dias, dict):
            return False
            
        for dia, valor in dias.items():
            if not isinstance(valor, bool) and valor is not None:
                return False
    
    return True

def calcular_dia_domingo(fecha_inicial_str):
    try:
        # Solo tomamos la parte de la fecha (días y meses)
        fecha_str = fecha_inicial_str.split()[0]  # Toma solo la parte de la fecha
        # Agregamos el año actual ya que solo viene día/mes
        año_actual = datetime.now().year
        fecha_completa = f"{fecha_str}/{año_actual}"  # Formato: DD/MM/AÑO
        fecha_inicial = datetime.strptime(fecha_completa, "%d/%m/%Y")  # Ajusta el formato
        
        # Calculamos días hasta el domingo
        dias_hasta_domingo = (6 - fecha_inicial.weekday()) % 7
        fecha_domingo = fecha_inicial + timedelta(days=dias_hasta_domingo)
        # Retornamos en el mismo formato DD/MM
        return fecha_domingo.strftime("%d/%m")
    except Exception as e:
        print(f"Error procesando la fecha: {e}")
        return None
    

def procesar_excel_dinamico(data):
    """
    Procesa la plantilla Excel y llena las celdas según la data recibida.
    """
    wb = load_workbook(get_template_path())
    worksheet = wb.active

    dias_columnas = {
        "lunes": ("E", "G"),
        "martes": ("H", "J"),
        "miercoles": ("K", "M"),
        "jueves": ("N", "P"),
        "viernes": ("Q", "S"),
        "sabado": ("T", "V"),
        "domingo": ("W", "Y")
    }

    estilo_formulario = {
        'font': Font(
            name='Arial',
            size=16,
            bold=True
        ),
        'alignment': Alignment(
            horizontal='center',
            vertical='center'
        )
    }


    def obtener_celda_principal(hoja, celda):
        """Obtiene la celda principal si está en un rango fusionado"""
        for merged_range in hoja.merged_cells.ranges:
            if celda.coordinate in merged_range:
                return hoja.cell(merged_range.min_row, merged_range.min_col)
        return celda
    
    print(f"Datos recibidos: {data}")

    if 'FORMULARIO' in data:
        formulario = data['FORMULARIO']
        campos_formulario = {
            'AÑO': 'I6',
            'PLACA': 'T7'
        }

        for campo, celda in campos_formulario.items():
            if campo in formulario:
                cell = worksheet[celda]
                cell.value = formulario[campo]
                cell.font = estilo_formulario['font']
                cell.alignment = estilo_formulario['alignment']
        if 'FECHA' in formulario:
            # Extraer solo la fecha sin la hora
            fecha_str = formulario['FECHA'].split()[0]
            # Asignar la fecha original a E6
            worksheet['E6'].value = fecha_str
            worksheet['E6'].font = estilo_formulario['font']
            worksheet['E6'].alignment = estilo_formulario['alignment']

            # Calcular y asignar el domingo a G6
            fecha_domingo = calcular_dia_domingo(formulario['FECHA'])
            worksheet['G6'].value = fecha_domingo
            worksheet['G6'].font = estilo_formulario['font']
            worksheet['G6'].alignment = estilo_formulario['alignment']

    inspeccion = data.get("INSPECCION", {})
    fila_inicial = 11
    
    for idx, (nombre_elemento, valores_dias) in enumerate(inspeccion.items()):
        fila_actual = fila_inicial + idx

        for dia, (col_inicio, col_fin) in dias_columnas.items():
            try:
                valor_dia = valores_dias.get(dia)
                celda_destino = worksheet[f"{col_inicio}{fila_actual}"]
                celda_principal = obtener_celda_principal(worksheet, celda_destino)

                if valor_dia is True:
                    celda_principal.value = "✔"
                    celda_principal.font = Font(name='Segoe UI Symbol', size=22, bold=True)
                    celda_principal.alignment = Alignment(horizontal='center', vertical='center')
                elif valor_dia is False:
                    celda_principal.value = "❌"
                    celda_principal.font = Font(name='Segoe UI Symbol', size=22, bold=True)
                    celda_principal.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    celda_principal.value = ""
                
            except Exception as e:
                print(f"Error procesando día {dia}: {str(e)}")

    if 'IMAGENES' in data:
        insertar_imagenes(worksheet, data['IMAGENES'])

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer

def insertar_imagenes(ws, imagenes_data):
    """Inserta las im��genes en el Excel"""
    celdas_imagenes = {
        'LOGO': 'B2',
        'FIRMA_USER': 'B25'
    }

    tamanos_fijos = {
        'LOGO': (200, 100),
        'FIRMA_USER': (180, 100)
    }

    modified_by = imagenes_data.get('MODIFICADO_POR', {})
    firmas_relevantes = imagenes_data.get('FIRMAS_RELV', {})

    print(f"Datos recibidos - MODIFICADO_POR: {modified_by}")
    print(f"Datos recibidos - FIRMAS_RELV: {firmas_relevantes}")
    
    grupos_firma_user = {
        'Lunes': ('E', 'G'),
        'Martes': ('H', 'J'),
        'Miercoles': ('K', 'M'),
        'Jueves': ('N', 'P'),
        'Viernes': ('Q', 'S'),
        'Sabado': ('T', 'V'),
        'Domingo': ('W', 'Y')
    }

    def verificar_contenido_columna(columna_inicio, columna_fin, fila_inicial=11, fila_final=20):
        for fila in range(fila_inicial, fila_final + 1):
            for col in range(ord(columna_inicio), ord(columna_fin) + 1):
                celda = ws[f"{chr(col)}{fila}"]
                if celda.value:
                    return True
        return False

    if 'LOGO' in imagenes_data:
        insertar_imagen_en_celda(ws, imagenes_data['LOGO'], 
                               celdas_imagenes['LOGO'], 
                               tamanos_fijos['LOGO'])

    for dia, (col_inicio, col_fin) in grupos_firma_user.items():
        print(f"\n=== Procesando firma para {dia} ===")
        tiene_contenido = verificar_contenido_columna(col_inicio, col_fin)
        print(f"¿Tiene contenido el día {dia}?: {tiene_contenido}")
        
        if tiene_contenido:
            col_media = chr(ord(col_inicio) + 1)
            celda_firma = f"{col_media}25"
            
            uid_modificador = modified_by.get(dia)
            print(f"UID del modificador para {dia}: {uid_modificador}")
            
            if uid_modificador:
                firma_key = f'FIRMA_USER_{uid_modificador}'
                print(f"Buscando firma con key: {firma_key}")
                print(f"Firmas disponibles: {firmas_relevantes.keys()}")
                
                if firma_key in firmas_relevantes:
                    firma_a_usar = firmas_relevantes[firma_key]
                    print(f"Insertando firma en celda {celda_firma}")
                    print(f"URL de la firma: {firma_a_usar}")
                    
                    insertar_imagen_en_celda(ws, firma_a_usar,
                                           celda_firma,
                                           tamanos_fijos['FIRMA_USER'])
                else:
                    print(f"No se encontró la firma para la key: {firma_key}")

def insertar_imagen_en_celda(ws, url, celda, tamano):
    """Inserta una imagen en una celda específica"""
    try:
        response = requests.get(url)
        response.raise_for_status()
        img_data = BytesIO(response.content)
        
        pil_image = Image.open(img_data)
        formato = pil_image.format.lower()
        
        formatos_soportados = ['png', 'jpeg', 'jpg']
        if formato not in formatos_soportados:
            print(f"Warning: Formato de imagen {formato} no soportado. Convirtiendo a PNG...")
            width_px, height_px = tamano
            pil_image = pil_image.convert('RGBA')
            pil_image = pil_image.resize((width_px, height_px), Image.LANCZOS)
            
            img_final = BytesIO()
            pil_image.save(img_final, format='PNG')
            img_final.seek(0)
        else:
            width_px, height_px = tamano
            pil_image = pil_image.resize((width_px, height_px), Image.LANCZOS)
            
            img_final = BytesIO()
            pil_image.save(img_final, format='PNG')
            img_final.seek(0)
        
        img = XLImage(img_final)
        img.width = width_px
        img.height = height_px
        
        ws.add_image(img, celda)
        print(f"Imagen insertada correctamente en la celda {celda}")
        
    except Exception as e:
        print(f"Error al insertar la imagen en la celda {celda}: {e}")