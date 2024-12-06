from io import BytesIO
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import os
import requests

# Construir la ruta relativa desde el directorio actual de ejecución
TEMPLATE_PATH = os.path.join(os.getcwd(), 'template', 'PREOPERACIONALES.xlsx')
OUTPUT_PATH = os.path.join(os.getcwd(), 'plantilla_modificada.xlsx')


def procesar_excel(data):
    # Verifica que el archivo de plantilla existe
    if not os.path.exists(TEMPLATE_PATH):
        print(f"ERROR: No se encontró el archivo de plantilla en {TEMPLATE_PATH}")
        raise FileNotFoundError("El archivo de plantilla de Excel no se encontró. Verifique la ruta.")

    # cargar el archivo de plantilla de Excel
    try:
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active
    except Exception as e:
        print(f"ERROR al cargar el archivo de plantilla: {str(e)}")
        raise e

    # Extraer el objeto FORMULARIO del JSON y eliminarlo del objeto original
    formulario_data = data.pop('FORMULARIO', None)
    pie_tabla = data.pop("PIE_TABLA", None)
    imagenes_data = data.pop("IMAGENES", {})

    # Llamar a la función que llena el formulario si existe
    if formulario_data:
        rellenar_formulario(ws, formulario_data)

    if pie_tabla:
        rellenar_pie_tabla(ws, pie_tabla)

    # Llamar a la función que llena la tabla en la sección específica
    rellenar_tabla(ws, data)
    
    if imagenes_data:
        insertar_imagenes(ws, imagenes_data, pie_tabla)

    

    # Guardar el archivo modificado
    wb.save(OUTPUT_PATH)


def obtener_rango_fusionado(hoja, celda):
    # Verifica si una celda está fusionada y obtiene la celda superior izquierda
    for merged_range in hoja.merged_cells.ranges:
        if celda.coordinate in merged_range:
            return merged_range, hoja.cell(merged_range.min_row, merged_range.min_col)
    return None, celda



def rellenar_formulario(ws, data):
    codigo = data.pop("Codigo", None)
    fecha_emision = data.pop("Fecha de Emision", None)
    km_total = data.pop("KM TOTAL", None)

    # Función para limpiar y normalizar el texto
    def normalizar_texto(texto):
        if isinstance(texto, str):  # Verifica si es una cadena de texto
            return texto.strip().lower().rstrip(':')
        return str(texto)  # Convierte otros tipos de datos a string

    # Función para actualizar el contenido de la celda sin cambiar la etiqueta y aplicar formato
    def actualizar_celda_con_etiqueta(celda, etiqueta, nuevo_valor):
        contenido_actual = str(celda.value)
        if etiqueta.lower() in contenido_actual.lower():
            # Mantener la etiqueta y reemplazar solo el valor
            partes = contenido_actual.split(':', 1)  # Separar por el primer ':'
            if len(partes) > 1:
                # Mantener la etiqueta original y aplicar negrita a la etiqueta
                nuevo_contenido = f"{partes[0].strip()}: {nuevo_valor}"
                celda.value = nuevo_contenido

                # Aplicar negrita a la etiqueta y normal al valor
                negrita_fuente = Font(bold=True)
                normal_fuente = Font(bold=False)

                # Aplicar estilos a la etiqueta y al valor
                celda.font = negrita_fuente  # Aplicar negrita a la celda completa (necesario para inicializar)
                partes_etiqueta = len(partes[0]) + 1  # Longitud de la etiqueta más ":"
                for i in range(partes_etiqueta, len(nuevo_contenido)):
                    celda.font = normal_fuente  # Cambiar el resto a fuente normal

            else:
                print(f"No se pudo actualizar la celda {celda.coordinate} correctamente.")
        else:
            print(f"No se encontró la etiqueta '{etiqueta}' en la celda {celda.coordinate}.")
    if km_total:
        # Obtener el rango fusionado y la celda principal para Q8
        celda_km = ws['Q8']
        rango_fusionado, celda_principal = obtener_rango_fusionado(ws, celda_km)
        
        # Aplicar el formato (Arial 14, negrita)
        fuente = Font(name='Arial', size=12, bold=True)
        celda_principal.font = fuente
        
        # Asignar el valor
        celda_principal.value = km_total

    # Iterar sobre los datos del formulario
    for key, value in data.items():
        # Normalizar la clave del JSON
        key_normalizado = normalizar_texto(key)
        found = False
        # Buscar la celda que coincide con la clave normalizada
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                # Normalizar el valor de la celda
                cell_value_normalizado = normalizar_texto(cell.value)
                # Si el valor de la celda coincide con la clave del formulario
                if cell_value_normalizado == key_normalizado:
                    # Obtener el rango fusionado y la celda principal
                    rango_fusionado, celda_principal = obtener_rango_fusionado(ws, cell)

                    if rango_fusionado:
                        # Determinar la celda a la derecha del rango fusionado
                        col_final = rango_fusionado.max_col + 1
                    else:
                        # Si no está fusionada, solo se mueve a la derecha
                        col_final = cell.column + 1

                    # Asignar el valor en la celda después del rango fusionado
                    ws.cell(row=cell.row, column=col_final).value = value

                    found = True
                    break
            if found:
                break
        if not found:
            print(f"No se encontró una celda para la clave '{key}'.")

def rellenar_tabla(ws, data):
    # Buscar la celda fusionada que contiene "item"
    fila_items = None
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col = merged_range.min_row, merged_range.min_col
        celda = ws.cell(row=min_row, column=min_col)
        if celda.value and str(celda.value).strip().lower() == "item":
            fila_items = min_row  # Guardar la fila donde se encuentra "item"
            break

    if not fila_items:
        raise ValueError("No se encontró la fila 'item' en la plantilla.")

    # Determinar dinámicamente las columnas de los días de la semana a partir de la fila "item"
    dias_columna = {}
    for col in range(1, ws.max_column + 1):
        celda = ws.cell(row=fila_items, column=col)
        dia = celda.value
        if dia and dia.strip().lower() in ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado', 'domingo']:
            columna_true = ws.cell(row=fila_items + 1, column=col).column_letter
            columna_false = ws.cell(row=fila_items + 1, column=col + 1).column_letter
            dias_columna[dia.strip().capitalize()] = (columna_true, columna_false)

    # Función para verificar si una celda está fusionada
    def obtener_celda_principal(hoja, celda):
        for merged_range in hoja.merged_cells.ranges:
            if celda.coordinate in merged_range:
                return hoja.cell(merged_range.min_row, merged_range.min_col)
        return celda

    # Iterar sobre las secciones del JSON
    for seccion, items in data.items():
        for item, dias in items.items():
            # Eliminar espacios en blanco y convertir a minúsculas los nombres de los items del JSON
            item_stripped_lower = item.strip().lower()

            # Encontrar la fila correspondiente al nombre del item
            for row in range(fila_items + 2, ws.max_row + 1):
                # Eliminar espacios en blanco y convertir a minúsculas los nombres de los items en la plantilla
                item_excel = ws[f'A{row}'].value
                if item_excel and item_excel.strip().lower() == item_stripped_lower:
                    # Rellenar los días de la semana en sus respectivas columnas
                    for dia, valor in dias.items():
                        dia_key = dia.strip().capitalize()
                        if dia_key in dias_columna:
                            columna_true, columna_false = dias_columna[dia_key]
                            if valor == True:
                                celda_destino_true = ws[f"{columna_true}{row}"]
                                celda_principal_true = obtener_celda_principal(ws, celda_destino_true)
                                celda_principal_true.value = 'X'
                            elif valor == False:
                                celda_destino_false = ws[f"{columna_false}{row}"]
                                celda_principal_false = obtener_celda_principal(ws, celda_destino_false)
                                celda_principal_false.value = 'X'
                    break


def rellenar_pie_tabla(ws, data):
    nombre_conductor = data.pop("Nombre del Conductor", None)
    observaciones = data.pop("OBSERVACIONES", None)

    # Llenar el nombre del conductor en la celda anterior
    if nombre_conductor:
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                # Normalizar el valor de la celda
                if cell.value and "nombre del conductor" in str(cell.value).strip().lower():
                    # Obtener el rango fusionado y la celda principal
                    rango_fusionado, celda_principal = obtener_rango_fusionado(ws, cell)

                    if rango_fusionado:
                        # Determinar la celda antes del rango fusionado
                        col_anterior = rango_fusionado.min_col - 1

                        if col_anterior >= 1:
                            celda_destino = ws.cell(row=cell.row, column=col_anterior)
                            celda_principal_destino = obtener_rango_fusionado(ws, celda_destino)[1]
                            celda_principal_destino.value = nombre_conductor

                    else:
                        # Si no está fusionada, simplemente toma la celda anterior
                        col_anterior = cell.column - 1
                        if col_anterior >= 1:
                            ws.cell(row=cell.row, column=col_anterior).value = nombre_conductor

                    found = True
                    break
            if found:
                break
        if not found:
            print("No se encontró una celda para 'Nombre del Conductor'.")

    # Llenar las observaciones en la celda debajo de la identificada
    if observaciones:
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                # Normalizar el valor de la celda
                if cell.value and "observaciones" in str(cell.value).strip().lower():
                    # Obtener el rango fusionado y la celda principal
                    rango_fusionado, celda_principal = obtener_rango_fusionado(ws, cell)

                    if rango_fusionado:
                        # Determinar la celda debajo del rango fusionado
                        fila_inferior = rango_fusionado.max_row + 1

                        if fila_inferior <= ws.max_row:
                            celda_destino = ws.cell(row=fila_inferior, column=cell.column)
                            celda_principal_destino = obtener_rango_fusionado(ws, celda_destino)[1]
                            celda_principal_destino.value = observaciones

                    else:
                        # Si no está fusionada, simplemente toma la celda de abajo
                        fila_inferior = cell.row + 1
                        if fila_inferior <= ws.max_row:
                            ws.cell(row=fila_inferior, column=cell.column).value = observaciones

                    found = True
                    break
            if found:
                break
        if not found:
            print("No se encontró una celda para 'OBSERVACIONES'.")

def insertar_imagenes(ws, imagenes_data, pie_tabla):
    celdas_imagenes = {
        'LOGO': 'A1',
        'FIRMA_USER': 'B84',
        'FIRMA_ENCARGADO': 'M84'
    }

    # Definir tamaños fijos para cada tipo de imagen (ancho, alto) en píxeles
    tamanos_fijos = {
        'FIRMA_REP': (110, 80),
        'LOGO': (200, 100), 
        'FIRMA_USER': (150, 75),  
        'FIRMA_ENCARGADO': (180, 105)
    }

    # Mapeo de columnas a días
    columnas_dias = {
        'H': 'Lunes',
        'J': 'Martes',
        'L': 'Miercoles',
        'N': 'Jueves',
        'P': 'Viernes',
        'R': 'Sabado',
        'T': 'Domingo'
    }

    # Grupos de celdas para verificar y celdas donde insertar firma
    grupos_celdas_firma_rep = [
        (['H14', 'H15', 'H16', 'H17', 'H18', 'H19', 'H21', 'H22', 'H23', 'H24', 'H52', 'H53', 'H54', 
          'I14', 'I15', 'I16', 'I17', 'I18', 'I19', 'I21', 'I22', 'I23', 'I24', 'I52', 'I53', 'I54'], 'H77'),
        (['J14', 'J15', 'J16', 'J17', 'J18', 'J19', 'J21', 'J22', 'J23', 'J24', 'J52', 'J53', 'J54', 
          'K14', 'K15', 'K16', 'K17', 'K18', 'K19', 'K21', 'K22', 'K23', 'K24', 'K52', 'K53', 'K54'], 'J77'),
        (['L14', 'L15', 'L16', 'L17', 'L18', 'L19', 'L21', 'L22', 'L23', 'L24', 'L52', 'L53', 'L54', 
          'M14', 'M15', 'M16', 'M17', 'M18', 'M19', 'M21', 'M22', 'M23', 'M24', 'M52', 'M53', 'M54'], 'L77'),
        (['N14', 'N15', 'N16', 'N17', 'N18', 'N19', 'N21', 'N22', 'N23', 'N24', 'N52', 'N53', 'N54', 
          'O14', 'O15', 'O16', 'O17', 'O18', 'O19', 'O21', 'O22', 'O23', 'O24', 'O52', 'O53', 'O54'], 'N77'),
        (['P14', 'P15', 'P16', 'P17', 'P18', 'P19', 'P21', 'P22', 'P23', 'P24', 'P52', 'P53', 'P54', 
          'Q14', 'Q15', 'Q16', 'Q17', 'Q18', 'Q19', 'Q21', 'Q22', 'Q23', 'Q24', 'Q52', 'Q53', 'Q54'], 'P77'),
        (['R14', 'R15', 'R16', 'R17', 'R18', 'R19', 'R21', 'R22', 'R23', 'R24', 'R52', 'R53', 'R54', 
          'S14', 'S15', 'S16', 'S17', 'S18', 'S19', 'S21', 'S22', 'S23', 'S24', 'S52', 'S53', 'S54'], 'R77'),
        (['T14', 'T15', 'T16', 'T17', 'T18', 'T19', 'T21', 'T22', 'T23', 'T24', 'T52', 'T53', 'T54', 
          'U14', 'U15', 'U16', 'U17', 'U18', 'U19', 'U21', 'U22', 'U23', 'U24', 'U52', 'U53', 'U54'], 'T77')
    ]

    # Obtener el diccionario modifiedBy del PIE_TABLA
    modified_by = pie_tabla.get('PIE_TABLA', {}).get('MODIFICADO_POR', {})
    print(f"DATA OBTENIDA PIE DE TABLA: {modified_by}")

    for tipo_imagen, url in imagenes_data.items():
        if url:
            if tipo_imagen in celdas_imagenes:
                celda = celdas_imagenes[tipo_imagen]
                insertar_imagen_en_celda(ws, url, celda, tamanos_fijos[tipo_imagen])
            elif tipo_imagen == 'FIRMA_REP' or tipo_imagen.startswith('FIRMA_USER_'):
                print("Insertando firmas en los días correspondientes...")
                for celdas_verificar, celda_firma in grupos_celdas_firma_rep:
                    if any(ws[celda].value for celda in celdas_verificar):
                        # Obtener el d��a correspondiente a este grupo de celdas
                        columna_dia = celda_firma[0]  
                        dia = columnas_dias.get(columna_dia)
                        print(f"\n=== Procesando día: {dia} ===")
                        print(f"Columna: {columna_dia}, Celda firma: {celda_firma}")
                        
                        # Obtener el UID del usuario que modificó ese día
                        uid_modificador = modified_by.get(dia)
                        print(f"UID del modificador para {dia}: {uid_modificador}")
                        
                        # Construir la clave de firma
                        firma_key = f'FIRMA_USER_{uid_modificador}'
                        print(f"Buscando firma con clave: {firma_key}")
                        
                        # Obtener el diccionario de firmas relevantes
                        firmas_relevantes = imagenes_data.get('FIRMAS_RELV', {})
                        print(f"Firmas relevantes disponibles: {list(firmas_relevantes.keys())}")
                        
                        # Verificar si existe una firma específica para ese UID en firmas_relevantes
                        if firma_key in firmas_relevantes:
                            firma_a_usar = firmas_relevantes[firma_key]
                            print(f"✅ Encontrada firma específica para {firma_key}")
                            print(f"URL de firma específica: {firma_a_usar[:50]}...")
                        else:
                            firma_a_usar = imagenes_data['FIRMA_USER']
                            print(f"❌ No se encontró firma específica para {firma_key}")
                            print(f"Usando firma por defecto (FIRMA_USER): {firma_a_usar[:50]}...")

                        if firma_a_usar:
                            print(f"Insertando firma en celda {celda_firma}")
                            insertar_imagen_en_celda(ws, firma_a_usar, celda_firma, tamanos_fijos['FIRMA_REP'])
                            print(f"✅ Firma insertada exitosamente en {celda_firma}")
                        else:
                            print(f"❌ Error: No se encontró firma para insertar en {celda_firma}")
                    else:
                        print(f"No se insertó firma en {celda_firma} porque no se encontró contenido en el grupo")

def insertar_imagen_en_celda(ws, url, celda, tamano):
    try:
        response = requests.get(url)
        response.raise_for_status()
        img_data = BytesIO(response.content)
        
        width_px, height_px = tamano
        
        pil_image = Image.open(img_data)
        pil_image = pil_image.resize((width_px, height_px), Image.LANCZOS)
        
        img_final = BytesIO()
        pil_image.save(img_final, format='PNG')
        img_final.seek(0)
        
        img = XLImage(img_final)
        img.width = width_px
        img.height = height_px
        
        ws.add_image(img, celda)
        print(f"Imagen insertada correctamente en la celda {celda}")
    except Exception as e:
        print(f"Error al insertar la imagen en la celda {celda}: {e}")

# ... (resto del código sin cambios)