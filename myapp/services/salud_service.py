import io
import os
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as XLImage
import requests
from io import BytesIO
import logging
from openpyxl.utils import get_column_letter, column_index_from_string

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def get_template_path():
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
    template_path = os.path.join(base_dir, 'template', 'AUTOREPORTE.xlsx')
    print(f"Template path: {template_path}")
    return template_path

def procesar_excel_salud(data):
    """
    Procesa la plantilla Excel y llena las celdas según la data recibida.
    """
    logger.info("==================== INICIO PROCESAMIENTO ====================")
    logger.info(f"Data recibida: {data}")
    
    try:
        wb = load_workbook(get_template_path())
        worksheet = wb.active
        print("Plantilla cargada exitosamente.")
    except Exception as e:
        logger.error(f"Error al cargar la plantilla de Excel: {e}")
        return None

    dias_columnas = {
        "lunes": ("AF", "AG"),
        "martes": ("AH", "AI"),
        "miercoles": ("AJ", "AK"),
        "jueves": ("AL", "AM"),
        "viernes": ("AN", "AO"),
        "sabado": ("AP", "AQ"),
        "domingo": ("AR", "AS")
    }

    # Estilo para las celdas del formulario
    estilo_formulario = {
        'font': Font(name='Arial', size=12, bold=True),
        'alignment': Alignment(horizontal='center', vertical='center')
    }

    def obtener_celda_principal(hoja, celda):
        """Obtiene la celda principal si está en un rango fusionado"""
        for merged_range in hoja.merged_cells.ranges:
            if celda.coordinate in merged_range:
                return hoja.cell(merged_range.min_row, merged_range.min_col)
        return celda

    # Procesar datos del formulario si existen
    if 'FORMULARIO' in data:
        formulario = data['FORMULARIO']
        campos_formulario = {
            'FECHA': 'D5:I5',
            'userName': 'J5:AE5',
            'cc': 'AG5:AN5',
            'rol': 'AO5:AS5',
            'contactoEmergencia': 'N7:AE7',
            'eps': 'D6:N6',
            'arl': 'S6:AE6',
            'afp': 'AG6:AK6',
            'proyecto': 'AO6:AS6',
            'telefonoEmergencia': 'AH7:AJ7',
            'parentesco': 'AM7',
            'direccion': 'AQ7:AR7'
        }
        for campo, celda in campos_formulario.items():
            if campo in formulario:
                try:
                    cell = worksheet[celda.split(':')[0]]
                    cell.value = formulario[campo]
                    cell.font = estilo_formulario['font']
                    cell.alignment = estilo_formulario['alignment']
                    print(f"Campo {campo} procesado en la celda {celda}.")
                except Exception as e:
                    logger.error(f"Error al procesar el campo {campo}: {e}")

    # Obtener la data de inspección
    inspeccion = data.get("PREGUNTAS", {})
    fila_inicial = 11
    
    # Iterar sobre los elementos en el JSON
    for idx, (nombre_elemento, valores_dias) in enumerate(inspeccion.items()):
        fila_actual = fila_inicial + idx
        print(f"Procesando elemento: {nombre_elemento}")

        # Iterar por cada día
        for dia, (col_inicio, col_fin) in dias_columnas.items():
            try:
                valor_dia = valores_dias.get(dia)
                print(f"Valor para {nombre_elemento} en {dia}: {valor_dia}")
                
                # Determinar la columna de destino
                if valor_dia is True:
                    celda_destino = worksheet[f"{col_inicio}{fila_actual}"]
                elif valor_dia is False:
                    celda_destino = worksheet[f"{col_fin}{fila_actual}"]
                else:
                    celda_destino = worksheet[f"{col_inicio}{fila_actual}"]
                
                celda_principal = obtener_celda_principal(worksheet, celda_destino)

                if valor_dia is True:
                    celda_principal.value = "✔"
                    celda_principal.font = Font(name='Segoe UI Symbol', size=22, bold=True)
                    celda_principal.alignment = Alignment(horizontal='center', vertical='center')
                    print(f"Celda {celda_destino.coordinate} actualizada con ✔")
                elif valor_dia is False:
                    celda_principal.value = "❌"
                    celda_principal.font = Font(name='Segoe UI Symbol', size=22, bold=True)
                    celda_principal.alignment = Alignment(horizontal='center', vertical='center')
                    print(f"Celda {celda_destino.coordinate} actualizada con ❌")
                else:
                    celda_principal.value = ""
                    print(f"Celda {celda_destino.coordinate} vacía")
                
            except Exception as e:
                print(f"Error en {nombre_elemento} - {dia}: {e}")

    logger.info("==================== FIN PROCESAMIENTO ====================")

    # Procesar imágenes si existen
    if 'IMAGENES' in data:
        try:
            insertar_imagenes_salud(worksheet, data['IMAGENES'])
        except Exception as e:
            logger.error(f"Error al insertar imágenes: {e}")

    try:
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        print("Archivo Excel guardado exitosamente.")
    except Exception as e:
        logger.error(f"Error al guardar el archivo Excel: {e}")
        return None

    return excel_buffer

def insertar_imagenes_salud(ws, imagenes_data):
    """Inserta las imágenes en el Excel"""
    # Configuración de celdas fijas
    celdas_imagenes = {
        'LOGO': 'A1',
        'TRANS': 'AO1'
    }

    # Tamaños fijos para cada tipo de imagen
    tamanos_fijos = {
        'LOGO': (250, 120),
        'FIRMA': (125,130)
    }

    # Grupos de celdas para verificar firma por día
    grupos_firma_user = {
        'lunes': ('AF', 'AG'),
        'martes': ('AH', 'AI'),
        'miércoles': ('AJ', 'AK'),
        'jueves': ('AL', 'AM'),
        'viernes': ('AN', 'AO'),
        'sábado': ('AP', 'AQ'),
        'domingo': ('AR', 'AS')
    }

    modified_by = imagenes_data.get('MODIFICADO_POR', {})
    firmas_relevantes = imagenes_data.get('FIRMAS_RELV', {})

    def verificar_contenido_columna(columna_inicio, columna_fin, fila_inicial=11, fila_final=13):
        """Verifica si hay contenido en el rango de celdas de una columna"""
        indice_inicio = column_index_from_string(columna_inicio)
        indice_fin = column_index_from_string(columna_fin)
        
        for fila in range(fila_inicial, fila_final + 1):
            for col in range(indice_inicio, indice_fin + 1):
                celda = ws[f"{get_column_letter(col)}{fila}"]
                if celda.value:
                    return True
        return False

    # Insertar logo si existe
    if 'LOGO' in imagenes_data:
        insertar_imagen_en_celda(ws, imagenes_data['LOGO'], 
                               celdas_imagenes['LOGO'], 
                               tamanos_fijos['LOGO'])
        print("Logo insertado.")

    # Insertar imágenes de firmas de usuario específicas por día
    if 'FIRMA_USER' in imagenes_data or 'FIRMAS_RELV' in imagenes_data:
        for dia, (col_inicio, col_fin) in grupos_firma_user.items():
            if verificar_contenido_columna(col_inicio, col_fin):
                # Calculamos la columna del medio
                indice_inicio = column_index_from_string(col_inicio)
                indice_fin = column_index_from_string(col_fin)
                indice_medio = (indice_inicio + indice_fin) // 2
                col_media = get_column_letter(indice_medio)
                
                celda_firma = f"{col_media}14"
                
                # Obtener el UID del usuario que modificó ese día
                uid_modificador = modified_by.get(dia)
                
                if uid_modificador:
                    # Construir la clave de firma específica
                    firma_key = f'FIRMA_USER_{uid_modificador}'
                    
                    # Determinar qué firma usar
                    if firma_key in firmas_relevantes:
                        # Usar la firma específica del usuario para ese día
                        firma_a_usar = firmas_relevantes[firma_key]
                        print(f"Usando firma específica para {dia}: {firma_key}")
                    else:
                        # Caer back a la firma por defecto si no hay firma específica
                        firma_a_usar = imagenes_data.get('FIRMA_USER')
                        print(f"Usando firma por defecto para {dia}")
                else:
                    # Si no hay uid modificador, usar la firma por defecto
                    firma_a_usar = imagenes_data.get('FIRMA_USER')
                    print(f"No se encontró UID modificador para {dia}, usando firma por defecto")
                
                # Insertar la firma si existe
                if firma_a_usar:
                    print(f"Insertando firma en la celda: {celda_firma} para el día: {dia}")
                    insertar_imagen_en_celda(ws, firma_a_usar,
                                           celda_firma,
                                           tamanos_fijos['FIRMA'])
                    print(f"Firma insertada para {dia}.")
                else:
                    print(f"No se encontró firma para {dia}")
            else:
                print(f"No se encontró contenido en las columnas {col_inicio}-{col_fin} para el día: {dia}")
def insertar_imagen_en_celda(ws, url, celda, tamano):
    """Inserta una imagen en una celda específica"""
    try:
        response = requests.get(url)
        response.raise_for_status()
        img_data = BytesIO(response.content)
        
        # Detectar el formato de la imagen
        pil_image = Image.open(img_data)
        formato = pil_image.format.lower()
        
        # Verificar formato soportado
        formatos_soportados = ['png', 'jpeg', 'jpg']
        if formato not in formatos_soportados:
            print(f"Warning: Formato de imagen {formato} no soportado. Convirtiendo a PNG...")
            # Convertir a PNG
            width_px, height_px = tamano
            pil_image = pil_image.convert('RGBA')
            pil_image = pil_image.resize((width_px, height_px), Image.LANCZOS)
            
            img_final = BytesIO()
            pil_image.save(img_final, format='PNG')
            img_final.seek(0)
        else:
            # Procesar normalmente
            width_px, height_px = tamano
            pil_image = pil_image.resize((width_px, height_px), Image.LANCZOS)
            
            img_final = BytesIO()
            pil_image.save(img_final, format='PNG')
            img_final.seek(0)
        
        # Crear y configurar la imagen para Excel
        img = XLImage(img_final)
        img.width = width_px
        img.height = height_px
        
        # Insertar la imagen
        ws.add_image(img, celda)
        print(f"Imagen insertada correctamente en la celda {celda}")
        
    except Exception as e:
        print(f"Error al insertar la imagen en la celda {celda}: {e}")
        # Continuar sin la imagen en caso de error
        pass